"""
评审管理 API 路由
"""
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, status, Query, Request
from sqlalchemy import func
from sqlalchemy.orm import Session
from app.core.database import get_db
from app.core.security import get_current_active_user, require_role, get_user_from_token
from app.models.user import User, UserRole
from app.models.team import Team, TeamMember
from app.models.work import Work, Review, WorkStatus
from app.models.setting import Log
from app.schemas.work import ReviewCreate, ReviewUpdate, ReviewResponse, WorkResponse
from app.schemas.common import PageResponse
from app.services.webhook import trigger_webhook_and_notification
from app.models.webhook import WebhookEventType
from app.models.setting import CompetitionTheme

router = APIRouter(prefix="/reviews", tags=["评审管理"])


def add_log(db: Session, user_id: int, action: str, resource: str = None,
            resource_id: int = None, details: str = None):
    """添加日志"""
    log = Log(user_id=user_id, action=action, resource=resource,
              resource_id=resource_id, details=details)
    db.add(log)
    db.commit()


@router.get("", response_model=PageResponse)
async def get_reviews(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    team_name: Optional[str] = None,
    work_name: Optional[str] = None,
    is_scored: Optional[bool] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.ADMIN, UserRole.REVIEWER))
):
    """获取评审列表（评审/管理员）"""
    query = db.query(Work).join(Team)

    if team_name:
        query = query.filter(Team.name.contains(team_name))

    if work_name:
        query = query.filter(Work.name.contains(work_name))

    if is_scored is not None:
        if is_scored:
            query = query.filter(Work.score != None)
        else:
            query = query.filter(Work.score == None)

    total = query.count()
    works = query.order_by(Work.created_at.desc()).offset((page - 1) * page_size).limit(page_size).all()

    # 计算最大点赞数（用于大众评分计算）
    max_vote_count = db.query(func.max(Work.vote_count)).scalar() or 1

    # 获取每个作品的评审信息
    items = []
    for work in works:
        review = db.query(Review).filter(
            Review.work_id == work.id,
            Review.user_id == current_user.id
        ).first()

        # 获取该作品的所有评审用于计算平均分
        all_reviews = db.query(Review).filter(Review.work_id == work.id).all()
        scored_reviews = [r for r in all_reviews if r.score is not None]
        avg_score = sum(r.score for r in scored_reviews) / len(scored_reviews) if scored_reviews else None

        # 计算大众评分得分：点赞数 / 最高点赞数 × 100
        public_score = (work.vote_count / max_vote_count * 100) if max_vote_count > 0 else 0

        # 计算最终得分：平均分×0.8 + 大众评分得分×0.2
        final_score = (avg_score * 0.8 + public_score * 0.2) if avg_score is not None else None

        work_data = WorkResponse.model_validate(work)
        work_data_dict = work_data.model_dump()
        work_data_dict["my_review"] = ReviewResponse.model_validate(review) if review else None
        work_data_dict["team_name"] = work.team.name
        work_data_dict["reviewer_name"] = current_user.nickname or current_user.username
        work_data_dict["score"] = avg_score
        work_data_dict["public_score"] = round(public_score, 1)
        work_data_dict["final_score"] = round(final_score, 1) if final_score is not None else None
        work_data_dict["max_vote_count"] = max_vote_count
        work_data_dict["review_count"] = len(all_reviews)
        items.append(work_data_dict)

    return PageResponse(
        total=total,
        page=page,
        page_size=page_size,
        items=items
    )


@router.post("", response_model=ReviewResponse)
async def create_review(
    review_data: ReviewCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.ADMIN, UserRole.REVIEWER))
):
    """提交评审"""
    work = db.query(Work).filter(Work.id == review_data.work_id).first()
    if not work:
        raise HTTPException(status_code=404, detail="作品不存在")

    # 检查是否已评审过
    existing = db.query(Review).filter(
        Review.work_id == review_data.work_id,
        Review.user_id == current_user.id
    ).first()

    if existing:
        raise HTTPException(status_code=400, detail="您已评审过该作品")

    # 创建评审
    review = Review(
        work_id=review_data.work_id,
        user_id=current_user.id,
        score=review_data.score,
        comment=review_data.comment
    )
    db.add(review)

    # 更新作品分数
    if review_data.score is not None:
        # 计算平均分
        all_reviews = db.query(Review).filter(Review.work_id == work.id).all()
        total_score = sum(r.score for r in all_reviews if r.score) + review_data.score
        work.score = total_score / (len(all_reviews) + 1)

    db.commit()
    db.refresh(review)

    add_log(db, current_user.id, "review", "work", work.id,
            f"评审作品: {work.name}, 分数: {review_data.score}")

    response = ReviewResponse(
        id=review.id,
        work_id=review.work_id,
        user_id=review.user_id,
        score=review.score,
        comment=review.comment,
        reviewer_name=current_user.nickname or current_user.username,
        created_at=review.created_at,
        updated_at=review.updated_at
    )

    # 触发 Webhook
    await trigger_webhook_and_notification(db, WebhookEventType.REVIEW_CREATED, {
        "id": review.id,
        "work_id": review.work_id,
        "work_name": work.name,
        "user_id": current_user.id,
        "reviewer_name": current_user.nickname or current_user.username,
        "score": review.score,
        "comment": review.comment
    }, "created")

    return response


@router.put("/{review_id:int}", response_model=ReviewResponse)
async def update_review(
    review_id: int,
    review_data: ReviewUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.ADMIN, UserRole.REVIEWER))
):
    """更新评审"""
    review = db.query(Review).filter(Review.id == review_id).first()
    if not review:
        raise HTTPException(status_code=404, detail="评审不存在")

    # 只有评审本人可以更新
    if review.user_id != current_user.id and current_user.role != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="权限不足")

    # 更新字段
    update_data = review_data.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(review, key, value)

    # 更新作品分数
    work = db.query(Work).filter(Work.id == review.work_id).first()
    if work and review.score is not None:
        all_reviews = db.query(Review).filter(Review.work_id == work.id).all()
        total_score = sum(r.score for r in all_reviews if r.score)
        work.score = total_score / len(all_reviews) if all_reviews else None

    db.commit()
    db.refresh(review)

    add_log(db, current_user.id, "update_review", "work", work.id,
            f"更新评审: {work.name}")

    response = ReviewResponse(
        id=review.id,
        work_id=review.work_id,
        user_id=review.user_id,
        score=review.score,
        comment=review.comment,
        reviewer_name=current_user.nickname or current_user.username,
        created_at=review.created_at,
        updated_at=review.updated_at
    )

    # 触发 Webhook
    await trigger_webhook_and_notification(db, WebhookEventType.REVIEW_UPDATED, {
        "id": review.id,
        "work_id": review.work_id,
        "work_name": work.name,
        "user_id": current_user.id,
        "score": review.score,
        "comment": review.comment,
        "updated_fields": list(update_data.keys())
    }, "updated")

    return response


@router.get("/my-reviews", response_model=PageResponse)
async def get_my_reviews(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.ADMIN, UserRole.REVIEWER))
):
    """获取我的评审列表"""
    query = db.query(Review).filter(Review.user_id == current_user.id)

    total = query.count()
    reviews = query.order_by(Review.created_at.desc()).offset((page - 1) * page_size).limit(page_size).all()

    items = []
    for review in reviews:
        work = db.query(Work).filter(Work.id == review.work_id).first()
        if work:
            items.append({
                "review": ReviewResponse.model_validate(review),
                "work": WorkResponse.model_validate(work)
            })

    return PageResponse(
        total=total,
        page=page,
        page_size=page_size,
        items=items
    )


@router.get("/all-by-work", response_model=PageResponse)
async def get_all_reviews_by_work(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    team_name: Optional[str] = None,
    work_name: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """获取所有评审（管理员）- 按作品分组显示"""
    query = db.query(Work).join(Team)

    if team_name:
        query = query.filter(Team.name.contains(team_name))

    if work_name:
        query = query.filter(Work.name.contains(work_name))

    total = query.count()
    works = query.order_by(Work.created_at.desc()).offset((page - 1) * page_size).limit(page_size).all()

    # 计算最大点赞数（用于大众评分计算）
    max_vote_count = db.query(func.max(Work.vote_count)).scalar() or 1

    items = []
    for work in works:
        # 获取该作品的所有评审
        reviews = db.query(Review).filter(Review.work_id == work.id).all()
        reviewers = []
        for r in reviews:
            reviewer = db.query(User).filter(User.id == r.user_id).first()
            reviewers.append({
                "review_id": r.id,
                "reviewer_name": reviewer.nickname or reviewer.username if reviewer else "未知",
                "score": r.score,
                "comment": r.comment,
                "created_at": r.created_at.isoformat() if r.created_at else None
            })

        # 计算平均分
        scored_reviews = [r for r in reviews if r.score is not None]
        avg_score = sum(r.score for r in scored_reviews) / len(scored_reviews) if scored_reviews else None

        # 计算大众评分得分：点赞数 / 最高点赞数 × 100
        public_score = (work.vote_count / max_vote_count * 100) if max_vote_count > 0 else 0

        # 计算最终得分：平均分×0.8 + 大众评分得分×0.2
        final_score = (avg_score * 0.8 + public_score * 0.2) if avg_score is not None else None

        items.append({
            "id": work.id,
            "work_id": work.id,
            "name": work.name,
            "team_id": work.team.id,
            "team_name": work.team.name,
            "theme_id": work.theme_id,
            "theme_name": work.theme_obj.name if work.theme_obj else None,
            "description": work.description,
            "agent_url": work.agent_url,
            "agent_editor_url": work.agent_editor_url,
            "pdf_file": work.pdf_file,
            "video_file": work.video_file,
            "status": work.status,
            "vote_count": work.vote_count,
            "max_vote_count": max_vote_count,
            "score": avg_score,
            "public_score": round(public_score, 1),
            "final_score": round(final_score, 1) if final_score is not None else None,
            "review_count": len(reviews),
            "reviews": reviewers
        })

    return PageResponse(
        total=total,
        page=page,
        page_size=page_size,
        items=items
    )


# ========== 成绩导出相关 ==========

def calculate_award(rank: int, total: int) -> str:
    """根据排名计算奖项"""
    if total == 0:
        return "参与奖"
    if rank == 1:
        return "特等奖"
    if rank <= 5:  # 1个特等奖 + 4个一等奖
        return "一等奖"
    if rank <= 15:  # 1+4+10
        return "二等奖"
    if rank <= 30:  # 1+4+10+15
        return "三等奖"
    return "参与奖"


@router.get("/export")
async def export_scores(
    request: Request,
    db: Session = Depends(get_db),
):
    """导出台成绩列表（仅管理员）- xlsx格式"""
    # 支持从 query token 或 header 认证
    auth_header = request.headers.get("Authorization") if request else None
    token = None
    if auth_header and auth_header.startswith("Bearer "):
        token = auth_header.replace("Bearer ", "")
    else:
        token = request.query_params.get("token") if request else None

    current_user = get_user_from_token(token, db) if token else None
    if not current_user:
        raise HTTPException(status_code=401, detail="无效的认证凭据")
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="仅管理员可导出")

    # 获取所有作品及其分数
    query = db.query(Work).join(Team).outerjoin(CompetitionTheme, Work.theme_id == CompetitionTheme.id)

    works = query.order_by(Work.created_at.desc()).all()

    # 计算最大点赞数（用于大众评分计算）
    max_vote_count = db.query(func.max(Work.vote_count)).scalar() or 1

    # 计算每件作品的分数
    work_scores = []
    for work in works:
        # 获取该作品的所有评审
        reviews = db.query(Review).filter(Review.work_id == work.id).all()
        scored_reviews = [r for r in reviews if r.score is not None]
        avg_score = sum(r.score for r in scored_reviews) / len(scored_reviews) if scored_reviews else None

        # 计算大众评分得分：点赞数 / 最高点赞数 × 100
        public_score = (work.vote_count / max_vote_count * 100) if max_vote_count > 0 else 0

        # 计算最终得分：平均分×0.8 + 大众评分得分×0.2
        final_score = (avg_score * 0.8 + public_score * 0.2) if avg_score is not None else None

        work_scores.append({
            "work": work,
            "avg_score": avg_score,
            "public_score": round(public_score, 1),
            "final_score": round(final_score, 1) if final_score is not None else None,
            "vote_count": work.vote_count,
            "team_members": work.team.members if hasattr(work.team, 'members') else []
        })

    # 按最终成绩从高到低排序
    work_scores.sort(key=lambda x: x["final_score"] if x["final_score"] is not None else 0, reverse=True)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from io import BytesIO

        wb = Workbook()
        ws = wb.active
        ws.title = "成绩列表"

        # 标题样式
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=11, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 写入表头
        headers = [
            "作品名称", "队伍", "队长姓名", "队长学工号",
            "队员姓名", "队员学工号", "主题", "投票数",
            "大众评分", "专业评审", "平均分", "最终成绩", "奖项"
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # 写入数据行
        total_works = len(work_scores)
        current_rank = 0
        prev_final_score = None

        for work_data in work_scores:
            work = work_data["work"]
            avg_score = work_data["avg_score"]
            public_score = work_data["public_score"]
            final_score = work_data["final_score"]
            vote_count = work_data["vote_count"]

            # 计算排名（相同分数并列）
            if final_score != prev_final_score:
                current_rank += 1
                prev_final_score = final_score

            award = calculate_award(current_rank, total_works)

            # 获取队长信息
            leader = db.query(User).filter(User.id == work.team.leader_id).first()
            leader_name = leader.nickname if leader else "-"
            leader_username = leader.username if leader else "-"

            # 获取队伍成员
            members = db.query(TeamMember).filter(TeamMember.team_id == work.team.id).all()
            theme_name = work.theme_obj.name if work.theme_obj else "-"

            # 每位队员一行
            for i, member in enumerate(members):
                member_user = db.query(User).filter(User.username == member.user_id).first()
                member_name = member_user.nickname if member_user else member.name
                member_username = member.user_id

                row_num = ws.max_row + 1

                # 判断是否是队长（队员姓名和队长姓名相同时使用队伍中的队长信息）
                is_leader_member = (i == 0)  # 第一个成员行输出队长信息

                row_data = [
                    work.name,                           # 作品名称
                    work.team.name,                      # 队伍
                    leader_name if is_leader_member else "",   # 队长姓名（仅首行）
                    leader_username if is_leader_member else "",  # 队长学工号（仅首行）
                    member_name,                         # 队员姓名
                    member_username,                     # 队员学工号
                    theme_name if i == 0 else "",        # 主题（仅首行）
                    vote_count if i == 0 else "",        # 投票数（仅首行）
                    public_score if i == 0 else "",      # 大众评分（仅首行）
                    f"{avg_score:.1f}" if avg_score and i == 0 else "",  # 专业评审（仅首行）
                    f"{avg_score:.1f}" if avg_score and i == 0 else "",  # 平均分（仅首行）
                    final_score if i == 0 and final_score else "",  # 最终成绩（仅首行）
                    award if i == 0 else ""              # 奖项（仅首行）
                ]

                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col, value=value)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        # 调整列宽
        column_widths = {
            'A': 20, 'B': 15, 'C': 12, 'D': 14,
            'E': 12, 'F': 14, 'G': 15, 'H': 8,
            'I': 10, 'J': 10, 'K': 10, 'L': 10, 'M': 10
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        from fastapi.responses import StreamingResponse
        import io

        return StreamingResponse(
            io.BytesIO(output.getvalue()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=scores.xlsx"}
        )
    except ImportError:
        raise HTTPException(status_code=500, detail="Excel导出功能需要安装 openpyxl")